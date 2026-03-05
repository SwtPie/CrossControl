import webview
import json
import sqlite3
import os
import sys
import re
import traceback
import logging
from datetime import datetime

# ── LOGGING ──────────────────────────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
LOG_PATH  = os.path.join(BASE_DIR, "cross_debug.log")
RACES_DIR = os.path.join(BASE_DIR, "races")
EVENTS_DB = os.path.join(BASE_DIR, "events.db")

os.makedirs(RACES_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
    ]
)
log = logging.getLogger("cross")
log.info(f"BASE_DIR  : {BASE_DIR}")
log.info(f"EVENTS_DB : {EVENTS_DB}")
log.info(f"RACES_DIR : {RACES_DIR}")

# ── CONNEXIONS ────────────────────────────────────────────────────────────────
_events_conn = None   # connexion à events.db
_event_conn  = None   # connexion à la DB de l'événement actif
_event_info  = None   # dict {id, nom, date, lieu} de l'événement actif

def get_events_db():
    global _events_conn
    if _events_conn is None:
        _events_conn = sqlite3.connect(EVENTS_DB, check_same_thread=False, timeout=10)
        _events_conn.row_factory = sqlite3.Row
        _events_conn.execute("PRAGMA journal_mode=DELETE")
        log.info("events.db connecté")
    return _events_conn

def get_event_db():
    if _event_conn is None:
        raise RuntimeError("Aucun événement ouvert")
    return _event_conn

def open_event_db(db_path):
    global _event_conn
    if _event_conn is not None:
        try:
            _event_conn.close()
        except Exception:
            pass
    _event_conn = sqlite3.connect(db_path, check_same_thread=False, timeout=10)
    _event_conn.row_factory = sqlite3.Row
    _event_conn.execute("PRAGMA journal_mode=DELETE")
    log.info(f"Event DB ouverte : {db_path}")
    return _event_conn

def parse_arg(data):
    if isinstance(data, str):
        return json.loads(data)
    if isinstance(data, dict):
        return data
    return data

def slugify(text):
    text = text.lower().strip()
    for src, dst in [("àâä","a"),("éèêë","e"),("îï","i"),("ôö","o"),("ùûü","u"),("ç","c")]:
        for c in src:
            text = text.replace(c, dst)
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text or "evenement"

# ── INIT EVENTS DB ────────────────────────────────────────────────────────────
def init_events_db():
    conn = get_events_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS evenements (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            nom         TEXT NOT NULL,
            date        TEXT,
            lieu        TEXT,
            description TEXT,
            slug        TEXT UNIQUE NOT NULL,
            db_path     TEXT NOT NULL,
            created_at  TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.commit()
    log.info("init_events_db() OK")

# ── INIT EVENT DB SCHEMA ──────────────────────────────────────────────────────
def init_event_db(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS participants (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            nom           TEXT NOT NULL,
            prenom        TEXT NOT NULL,
            classe        TEXT,
            etablissement TEXT,
            sexe          TEXT,
            vma           REAL,
            dossard       INTEGER UNIQUE,
            created_at    TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS courses (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            nom         TEXT NOT NULL,
            distance    REAL,
            vma_min     REAL,
            vma_max     REAL,
            statut      TEXT DEFAULT 'preparation',
            started_at  TEXT,
            finished_at TEXT,
            masquee     INTEGER DEFAULT 0,
            created_at  TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS course_participants (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            course_id      INTEGER,
            participant_id INTEGER,
            FOREIGN KEY(course_id)      REFERENCES courses(id),
            FOREIGN KEY(participant_id) REFERENCES participants(id)
        );
        CREATE TABLE IF NOT EXISTS arrivees (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            course_id      INTEGER,
            participant_id INTEGER,
            ordre_arrivee  INTEGER,
            temps_secondes REAL,
            dossard_saisi  INTEGER,
            FOREIGN KEY(course_id)      REFERENCES courses(id),
            FOREIGN KEY(participant_id) REFERENCES participants(id)
        );
    """)
    conn.commit()
    try:
        conn.execute("ALTER TABLE courses ADD COLUMN masquee INTEGER DEFAULT 0")
        conn.commit()
    except Exception:
        pass


# ════════════════════════════════════════════════════════════════════════════════
# API
# ════════════════════════════════════════════════════════════════════════════════
class API:

    # ─── GESTION DES ÉVÉNEMENTS ──────────────────────────────────────────────────

    def get_evenements(self):
        log.debug("API.get_evenements()")
        try:
            conn = get_events_db()
            rows = conn.execute("SELECT * FROM evenements ORDER BY created_at DESC").fetchall()
            result = []
            for r in rows:
                e = dict(r)
                try:
                    ec = sqlite3.connect(e['db_path'])
                    ec.row_factory = sqlite3.Row
                    e['nb_participants'] = ec.execute("SELECT COUNT(*) as c FROM participants").fetchone()['c']
                    e['nb_courses']      = ec.execute("SELECT COUNT(*) as c FROM courses").fetchone()['c']
                    ec.close()
                except Exception:
                    e['nb_participants'] = 0
                    e['nb_courses']      = 0
                result.append(e)
            return result
        except Exception as e:
            log.error(f"get_evenements ERREUR: {e}\n{traceback.format_exc()}")
            return []

    def create_evenement(self, data):
        log.info(f"API.create_evenement()")
        try:
            data = parse_arg(data)
            nom  = data.get('nom', '').strip()
            if not nom:
                return {"success": False, "error": "Nom requis"}
            conn      = get_events_db()
            base_slug = slugify(nom)
            slug, i   = base_slug, 1
            while conn.execute("SELECT id FROM evenements WHERE slug=?", (slug,)).fetchone():
                slug = f"{base_slug}_{i}"; i += 1
            db_path = os.path.join(RACES_DIR, f"{slug}.db")
            ec = sqlite3.connect(db_path)
            init_event_db(ec)
            ec.close()
            conn.execute(
                "INSERT INTO evenements (nom, date, lieu, description, slug, db_path) VALUES (?,?,?,?,?,?)",
                (nom, data.get('date',''), data.get('lieu',''), data.get('description',''), slug, db_path)
            )
            conn.commit()
            log.info(f"  → événement créé : {slug}")
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}

    def update_evenement(self, eid, data):
        log.info(f"API.update_evenement(eid={eid})")
        try:
            data = parse_arg(data)
            conn = get_events_db()
            conn.execute(
                "UPDATE evenements SET nom=?, date=?, lieu=?, description=? WHERE id=?",
                (data['nom'], data.get('date',''), data.get('lieu',''), data.get('description',''), eid)
            )
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}

    def delete_evenement(self, eid):
        log.info(f"API.delete_evenement(eid={eid})")
        try:
            conn = get_events_db()
            row  = conn.execute("SELECT db_path FROM evenements WHERE id=?", (eid,)).fetchone()
            if row:
                try:
                    os.remove(row['db_path'])
                except FileNotFoundError:
                    pass
            conn.execute("DELETE FROM evenements WHERE id=?", (eid,))
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}

    def open_evenement(self, eid):
        global _event_info
        log.info(f"API.open_evenement(eid={eid})")
        try:
            conn = get_events_db()
            row  = conn.execute("SELECT * FROM evenements WHERE id=?", (eid,)).fetchone()
            if not row:
                return {"success": False, "error": "Événement introuvable"}
            ec = open_event_db(row['db_path'])
            init_event_db(ec)
            _event_info = {"id": row['id'], "nom": row['nom'], "date": row['date'], "lieu": row['lieu']}
            log.info(f"  → ouvert : {row['nom']}")
            return {"success": True, "evenement": _event_info}
        except Exception as e:
            log.error(f"  ERREUR: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}

    def close_evenement(self):
        global _event_conn, _event_info
        log.info("API.close_evenement()")
        if _event_conn:
            try:
                _event_conn.close()
            except Exception:
                pass
            _event_conn = None
        _event_info = None
        return {"success": True}

    def get_current_evenement(self):
        return _event_info  # None si aucun événement ouvert

    # ─── PARTICIPANTS ────────────────────────────────────────────────────────────

    def get_participants(self):
        try:
            rows = get_event_db().execute("SELECT * FROM participants ORDER BY nom, prenom").fetchall()
            return [dict(r) for r in rows]
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return []

    def add_participant(self, data):
        log.info(f"API.add_participant() val={data}")
        try:
            data = parse_arg(data)
            get_event_db().execute(
                "INSERT INTO participants (nom, prenom, classe, etablissement, sexe, vma, dossard) VALUES (?,?,?,?,?,?,?)",
                (data['nom'], data['prenom'], data.get('classe',''), data.get('etablissement',''),
                 data.get('sexe',''), data.get('vma'), data.get('dossard'))
            )
            get_event_db().commit()
            return {"success": True}
        except sqlite3.IntegrityError as e:
            if "participants.dossard" in str(e):
                return {"success": False, "error": f"Le dossard {data.get('dossard')} est déjà attribué."}
            return {"success": False, "error": str(e)}
        except Exception as e:
            log.error(f"  ERREUR: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}

    def update_participant(self, pid, data):
        try:
            data = parse_arg(data)
            get_event_db().execute(
                "UPDATE participants SET nom=?, prenom=?, classe=?, etablissement=?, sexe=?, vma=?, dossard=? WHERE id=?",
                (data['nom'], data['prenom'], data.get('classe',''), data.get('etablissement',''),
                 data.get('sexe',''), data.get('vma'), data.get('dossard'), pid)
            )
            get_event_db().commit()
            return {"success": True}
        except sqlite3.IntegrityError as e:
            if "participants.dossard" in str(e):
                return {"success": False, "error": f"Le dossard {data.get('dossard')} est déjà attribué."}
            return {"success": False, "error": str(e)}
        except Exception as e:
            log.error(f"  ERREUR: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}

    def delete_participant(self, pid):
        try:
            get_event_db().execute("DELETE FROM participants WHERE id=?", (pid,))
            get_event_db().commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def auto_assign_dossards(self, start=1):
        try:
            conn = get_event_db()
            # Grouper par établissement (ordre alphabétique), tri nom/prénom dans chaque groupe
            # Participants sans établissement regroupés en dernier
            rows = conn.execute("""
                SELECT id, etablissement FROM participants
                ORDER BY
                    CASE WHEN etablissement IS NULL OR etablissement = '' THEN 1 ELSE 0 END,
                    etablissement COLLATE NOCASE,
                    nom COLLATE NOCASE,
                    prenom COLLATE NOCASE
            """).fetchall()

            GAP = 5  # numéros de marge entre chaque établissement
            current      = int(start)
            current_etab = None
            count        = 0

            for p in rows:
                etab = p["etablissement"] or ""
                if current_etab is None:
                    current_etab = etab
                elif etab != current_etab:
                    # Prochain multiple de GAP + GAP (ex: après 23 → 30)
                    current = (((current - 1) // GAP) + 2) * GAP + 1
                    current_etab = etab
                conn.execute("UPDATE participants SET dossard=? WHERE id=?", (current, p["id"]))
                current += 1
                count   += 1

            conn.commit()
            return {"success": True, "count": count}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    # ─── COURSES ─────────────────────────────────────────────────────────────────

    def get_courses(self):
        try:
            rows = get_event_db().execute("SELECT * FROM courses ORDER BY created_at DESC").fetchall()
            return [dict(r) for r in rows]
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return []

    def add_course(self, data):
        try:
            data = parse_arg(data)
            conn = get_event_db()
            c = conn.execute(
                "INSERT INTO courses (nom, distance, vma_min, vma_max) VALUES (?,?,?,?)",
                (data['nom'], data.get('distance'), data.get('vma_min'), data.get('vma_max'))
            )
            conn.commit()
            return {"success": True, "id": c.lastrowid}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def update_course(self, cid, data):
        try:
            data = parse_arg(data)
            conn = get_event_db()
            conn.execute(
                "UPDATE courses SET nom=?, distance=?, vma_min=?, vma_max=? WHERE id=?",
                (data['nom'], data.get('distance'), data.get('vma_min'), data.get('vma_max'), cid)
            )
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def delete_course(self, cid):
        try:
            conn = get_event_db()
            conn.execute("DELETE FROM course_participants WHERE course_id=?", (cid,))
            conn.execute("DELETE FROM arrivees WHERE course_id=?", (cid,))
            conn.execute("DELETE FROM courses WHERE id=?", (cid,))
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def get_course_participants(self, course_id):
        try:
            rows = get_event_db().execute("""
                SELECT p.* FROM participants p
                JOIN course_participants cp ON cp.participant_id = p.id
                WHERE cp.course_id = ? ORDER BY p.dossard, p.nom
            """, (course_id,)).fetchall()
            return [dict(r) for r in rows]
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return []

    def add_participant_to_course(self, course_id, participant_id):
        try:
            conn = get_event_db()
            if not conn.execute("SELECT id FROM course_participants WHERE course_id=? AND participant_id=?",
                                (course_id, participant_id)).fetchone():
                conn.execute("INSERT INTO course_participants (course_id, participant_id) VALUES (?,?)",
                             (course_id, participant_id))
                conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def remove_participant_from_course(self, course_id, participant_id):
        try:
            conn = get_event_db()
            conn.execute("DELETE FROM course_participants WHERE course_id=? AND participant_id=?",
                         (course_id, participant_id))
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def auto_add_by_vma(self, course_id):
        try:
            conn   = get_event_db()
            course = conn.execute("SELECT * FROM courses WHERE id=?", (course_id,)).fetchone()
            if not course:
                return {"success": False, "error": "Course introuvable"}
            if course['vma_min'] is None or course['vma_max'] is None:
                return {"success": False, "error": "VMA min/max non définis"}
            parts = conn.execute("SELECT id FROM participants WHERE vma >= ? AND vma <= ?",
                                 (course['vma_min'], course['vma_max'])).fetchall()
            count = 0
            for p in parts:
                if not conn.execute("SELECT id FROM course_participants WHERE course_id=? AND participant_id=?",
                                    (course_id, p['id'])).fetchone():
                    conn.execute("INSERT INTO course_participants (course_id, participant_id) VALUES (?,?)",
                                 (course_id, p['id']))
                    count += 1
            conn.commit()
            return {"success": True, "count": count}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def toggle_masquer_course(self, cid):
        try:
            conn    = get_event_db()
            current = conn.execute("SELECT masquee FROM courses WHERE id=?", (cid,)).fetchone()
            if not current:
                return {"success": False, "error": "Course introuvable"}
            new_val = 0 if current["masquee"] else 1
            conn.execute("UPDATE courses SET masquee=? WHERE id=?", (new_val, cid))
            conn.commit()
            return {"success": True, "masquee": new_val}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def get_courses_terminees(self):
        try:
            rows = get_event_db().execute(
                "SELECT * FROM courses WHERE statut='terminee' ORDER BY finished_at DESC"
            ).fetchall()
            return [dict(r) for r in rows]
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return []

    # ─── COURSE EN DIRECT ────────────────────────────────────────────────────────

    def start_course(self, course_id):
        try:
            conn = get_event_db()
            now  = datetime.now().isoformat()
            conn.execute("UPDATE courses SET statut='en_cours', started_at=? WHERE id=?", (now, course_id))
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def finish_course(self, course_id):
        try:
            conn = get_event_db()
            conn.execute("UPDATE courses SET statut='terminee', finished_at=? WHERE id=?",
                         (datetime.now().isoformat(), course_id))
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def reset_course(self, course_id):
        try:
            conn = get_event_db()
            conn.execute("DELETE FROM arrivees WHERE course_id=?", (course_id,))
            conn.execute("UPDATE courses SET statut='preparation', started_at=NULL, finished_at=NULL WHERE id=?",
                         (course_id,))
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def enregistrer_arrivee(self, course_id, temps_secondes):
        try:
            conn  = get_event_db()
            ordre = conn.execute("SELECT COUNT(*) as cnt FROM arrivees WHERE course_id=?",
                                 (course_id,)).fetchone()['cnt'] + 1
            conn.execute("INSERT INTO arrivees (course_id, ordre_arrivee, temps_secondes) VALUES (?,?,?)",
                         (course_id, ordre, temps_secondes))
            conn.commit()
            aid = conn.execute("SELECT id FROM arrivees WHERE course_id=? AND ordre_arrivee=?",
                               (course_id, ordre)).fetchone()['id']
            return {"success": True, "ordre": ordre, "arrivee_id": aid}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def assigner_dossard_arrivee(self, arrivee_id, dossard):
        try:
            conn = get_event_db()
            p    = conn.execute("SELECT id FROM participants WHERE dossard=?", (dossard,)).fetchone()
            if not p:
                return {"success": False, "error": f"Dossard {dossard} introuvable"}
            conn.execute("UPDATE arrivees SET dossard_saisi=?, participant_id=? WHERE id=?",
                         (dossard, p['id'], arrivee_id))
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def get_arrivees(self, course_id):
        try:
            rows = get_event_db().execute("""
                SELECT a.*, p.nom, p.prenom, p.classe, p.etablissement, p.sexe, p.vma,
                       p.dossard as dossard_participant
                FROM arrivees a
                LEFT JOIN participants p ON p.id = a.participant_id
                WHERE a.course_id = ? ORDER BY a.ordre_arrivee
            """, (course_id,)).fetchall()
            return [dict(r) for r in rows]
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return []

    def supprimer_arrivee(self, arrivee_id):
        try:
            conn = get_event_db()
            a    = conn.execute("SELECT * FROM arrivees WHERE id=?", (arrivee_id,)).fetchone()
            if not a:
                return {"success": False, "error": "Arrivée introuvable"}
            conn.execute("DELETE FROM arrivees WHERE id=?", (arrivee_id,))
            conn.execute("UPDATE arrivees SET ordre_arrivee=ordre_arrivee-1 WHERE course_id=? AND ordre_arrivee>?",
                         (a['course_id'], a['ordre_arrivee']))
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def ajouter_arrivee_manuelle(self, course_id, temps_secondes, dossard):
        try:
            conn = get_event_db()
            p    = conn.execute("SELECT id FROM participants WHERE dossard=?", (dossard,)).fetchone()
            if not p:
                return {"success": False, "error": f"Dossard {dossard} introuvable"}
            ordre = conn.execute("SELECT COUNT(*) as cnt FROM arrivees WHERE course_id=?",
                                 (course_id,)).fetchone()['cnt'] + 1
            conn.execute("INSERT INTO arrivees (course_id, ordre_arrivee, temps_secondes, dossard_saisi, participant_id) VALUES (?,?,?,?,?)",
                         (course_id, ordre, temps_secondes, dossard, p['id']))
            conn.commit()
            return {"success": True}
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return {"success": False, "error": str(e)}

    def get_classement(self, course_id):
        try:
            conn   = get_event_db()
            course = conn.execute("SELECT * FROM courses WHERE id=?", (course_id,)).fetchone()
            if not course:
                return []
            rows = conn.execute("""
                SELECT a.*, p.nom, p.prenom, p.classe, p.etablissement, p.sexe, p.vma,
                       p.dossard as num_dossard
                FROM arrivees a
                LEFT JOIN participants p ON p.id = a.participant_id
                WHERE a.course_id=? AND a.participant_id IS NOT NULL ORDER BY a.ordre_arrivee
            """, (course_id,)).fetchall()
            result = []
            for r in rows:
                d = dict(r)
                if d['temps_secondes'] and d['temps_secondes'] > 0 and course['distance']:
                    v = (course['distance'] / d['temps_secondes']) * 3.6
                    d['vitesse_kmh'] = round(v, 2)
                    d['pct_vma']     = round((v / d['vma']) * 100, 1) if d['vma'] else None
                else:
                    d['vitesse_kmh'] = None
                    d['pct_vma']     = None
                result.append(d)
            return result
        except Exception as e:
            log.error(f"  ERREUR: {e}"); return []


    # ─── IMPORT PARTICIPANTS ──────────────────────────────────────────────────────

    def parse_import_file(self, file_b64, filename, etablissement_override=""):
        """
        Reçoit un fichier CSV ou XLSX encodé en base64.
        Retourne un aperçu : liste de lignes avec statut (ok / warning / doublon / erreur).
        """
        import base64, io, csv
        log.info(f"API.parse_import_file() filename={filename}")
        try:
            raw = base64.b64decode(file_b64)
            ext = filename.rsplit(".", 1)[-1].lower()

            rows_raw = []

            if ext == "csv":
                # Détecter le séparateur
                sample = raw[:2048].decode("utf-8-sig", errors="replace")
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                reader = csv.DictReader(
                    io.StringIO(raw.decode("utf-8-sig", errors="replace")),
                    dialect=dialect
                )
                for r in reader:
                    rows_raw.append({k.strip().lower(): v.strip() for k, v in r.items()})

            elif ext in ("xlsx", "xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                    ws = wb.active
                    rows_iter = iter(ws.iter_rows(values_only=True))
                    headers = [str(c).strip().lower() if c else "" for c in next(rows_iter)]
                    for row in rows_iter:
                        rows_raw.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
                except ImportError:
                    return {"success": False, "error": "openpyxl non installé (pip install openpyxl)"}
            else:
                return {"success": False, "error": f"Format non supporté : .{ext}"}

            # Mapping flexible des colonnes
            ALIASES = {
                "nom":           ["nom", "name", "last_name", "lastname", "family_name"],
                "prenom":        ["prenom", "prénom", "first_name", "firstname", "given_name"],
                "classe":        ["classe", "class", "group", "groupe", "level", "niveau"],
                "etablissement": ["etablissement", "établissement", "school", "ecole", "école", "structure"],
                "sexe":          ["sexe", "genre", "sex", "gender", "m/f"],
                "vma":           ["vma", "vma (km/h)", "vitesse", "speed"],
                "dossard":       ["dossard", "bib", "numero", "numéro", "n°", "num"],
            }

            def map_col(row, field):
                for alias in ALIASES[field]:
                    for k in row:
                        if k == alias or k.strip("'\"\t ") == alias:
                            return row[k]
                return ""

            def normalize_name(s):
                """Title case + gestion des noms composés"""
                if not s:
                    return ""
                return " ".join(w.capitalize() for w in s.strip().split())

            def normalize_sexe(s):
                s = s.strip().upper()
                if s in ("F", "FEMININ", "FÉMININ", "FEMALE", "FILLE"):
                    return "F"
                if s in ("M", "MASCULIN", "MALE", "GARCON", "GARÇON"):
                    return "M"
                return ""

            # Participants déjà en base pour détection des doublons
            conn = get_event_db()
            existing = conn.execute("SELECT nom, prenom, classe FROM participants").fetchall()
            existing_set = {(r["nom"].lower(), r["prenom"].lower()) for r in existing}

            result = []
            for i, raw_row in enumerate(rows_raw):
                nom    = normalize_name(map_col(raw_row, "nom"))
                prenom = normalize_name(map_col(raw_row, "prenom"))

                if not nom and not prenom:
                    continue  # ligne vide, on ignore

                warnings = []
                statut   = "ok"

                if not nom:
                    statut = "erreur"
                    warnings.append("Nom manquant")
                if not prenom:
                    statut = "erreur"
                    warnings.append("Prénom manquant")

                # Doublon
                if nom and prenom and (nom.lower(), prenom.lower()) in existing_set:
                    statut = "doublon"
                    warnings.append("Déjà présent dans l'événement")

                # Établissement
                etab = etablissement_override.strip() if etablissement_override.strip() else normalize_name(map_col(raw_row, "etablissement"))

                # VMA
                vma_raw = map_col(raw_row, "vma").replace(",", ".")
                vma = None
                try:
                    vma = float(vma_raw) if vma_raw else None
                except ValueError:
                    warnings.append("VMA invalide ignorée")
                if vma is None and statut == "ok":
                    statut = "warning"
                    warnings.append("VMA manquante")

                # Sexe
                sexe = normalize_sexe(map_col(raw_row, "sexe"))
                if not sexe and statut == "ok":
                    statut = "warning"
                    warnings.append("Sexe non renseigné")

                # Dossard
                dossard_raw = map_col(raw_row, "dossard")
                dossard = None
                try:
                    dossard = int(dossard_raw) if dossard_raw else None
                except ValueError:
                    pass

                # Classe
                classe = map_col(raw_row, "classe").strip()

                result.append({
                    "index":         i,
                    "nom":           nom,
                    "prenom":        prenom,
                    "classe":        classe,
                    "etablissement": etab,
                    "sexe":          sexe,
                    "vma":           vma,
                    "dossard":       dossard,
                    "statut":        statut,
                    "warnings":      warnings,
                })

            counts = {
                "total":    len(result),
                "ok":       sum(1 for r in result if r["statut"] == "ok"),
                "warning":  sum(1 for r in result if r["statut"] == "warning"),
                "doublon":  sum(1 for r in result if r["statut"] == "doublon"),
                "erreur":   sum(1 for r in result if r["statut"] == "erreur"),
            }

            return {"success": True, "rows": result, "counts": counts}

        except Exception as e:
            log.error(f"parse_import_file ERREUR: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}

    def confirm_import(self, rows, skip_doublons=True):
        """
        Insère en base les lignes validées (statut != erreur, doublons selon option).
        """
        log.info(f"API.confirm_import() nb_rows={len(rows)}")
        try:
            rows = parse_arg(rows) if isinstance(rows, str) else rows
            conn = get_event_db()
            imported = 0
            skipped  = 0
            errors   = []

            for r in rows:
                statut = r.get("statut", "ok")
                if statut == "erreur":
                    skipped += 1
                    continue
                if statut == "doublon" and skip_doublons:
                    skipped += 1
                    continue
                try:
                    conn.execute(
                        "INSERT INTO participants (nom, prenom, classe, etablissement, sexe, vma, dossard) VALUES (?,?,?,?,?,?,?)",
                        (r["nom"], r["prenom"], r.get("classe",""), r.get("etablissement",""),
                         r.get("sexe",""), r.get("vma"), r.get("dossard"))
                    )
                    imported += 1
                except sqlite3.IntegrityError:
                    errors.append(f"Dossard {r.get('dossard')} déjà attribué — {r['nom']} {r['prenom']} ignoré")
                    skipped += 1

            conn.commit()
            return {"success": True, "imported": imported, "skipped": skipped, "errors": errors}
        except Exception as e:
            log.error(f"confirm_import ERREUR: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}


        # ─── EXPORT ──────────────────────────────────────────────────────────────────

    def export_file(self, params_json):
        """
        Export unifié : context=(participants|course), format=(pdf|csv|xlsx),
        filters={sexe, etablissement, classe}
        """
        import json as _json
        log.info(f"API.export_file() params={params_json}")
        try:
            params  = _json.loads(params_json) if isinstance(params_json, str) else params_json
            context = params.get("context", "participants")
            fmt     = params.get("format", "pdf")
            filters = params.get("filters", {})
            course_id = params.get("course_id")

            conn  = get_event_db()
            event = _event_info or {}

            # ── Récupérer les lignes selon le contexte ──
            if context == "course" and course_id:
                course = conn.execute("SELECT * FROM courses WHERE id=?", (course_id,)).fetchone()
                rows = conn.execute("""
                    SELECT p.* FROM participants p
                    JOIN course_participants cp ON cp.participant_id = p.id
                    WHERE cp.course_id = ? ORDER BY p.dossard, p.nom
                """, (course_id,)).fetchall()
                title_line = dict(course)["nom"] if course else "Course"
            else:
                course = None
                rows = conn.execute("SELECT * FROM participants ORDER BY dossard, nom, prenom").fetchall()
                title_line = "Tous les participants"

            rows = [dict(r) for r in rows]

            # ── Appliquer les filtres (ET) ──
            sexe  = filters.get("sexe", "")
            etab  = filters.get("etablissement", "")
            cls   = filters.get("classe", "")
            if sexe:  rows = [r for r in rows if r.get("sexe") == sexe]
            if etab:  rows = [r for r in rows if r.get("etablissement") == etab]
            if cls:   rows = [r for r in rows if r.get("classe") == cls]

            # ── Dossier exports ──
            exports_dir = os.path.join(BASE_DIR, "exports")
            os.makedirs(exports_dir, exist_ok=True)

            safe_title = title_line.replace(" ", "_")
            filter_tag = "_".join(filter(None, [sexe, etab, cls])).replace(" ", "_")
            if filter_tag: safe_title += f"_{filter_tag}"

            # ════════════════════════════════
            if fmt == "csv":
                import csv, io
                path = os.path.join(exports_dir, f"{safe_title}.csv")
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.DictWriter(f, fieldnames=["dossard","nom","prenom","classe","etablissement","sexe","vma"],
                                           extrasaction="ignore")
                    writer.writeheader()
                    for r in rows:
                        writer.writerow({k: (r.get(k) or "") for k in ["dossard","nom","prenom","classe","etablissement","sexe","vma"]})

            # ════════════════════════════════
            elif fmt == "xlsx":
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = title_line[:31]

                accent  = "FFD60A3C"
                dark    = "FF161920"
                light   = "FFF7F7F7"
                border_color = "FFD0D5E2"

                thin = Border(
                    left=Side(style="thin", color=border_color),
                    right=Side(style="thin", color=border_color),
                    top=Side(style="thin", color=border_color),
                    bottom=Side(style="thin", color=border_color),
                )

                # Titre événement
                ws.merge_cells("A1:G1")
                title_cell = ws["A1"]
                title_cell.value = f"CROSSCONTROL — {event.get('nom','')} — {title_line}"
                title_cell.font = Font(name="Arial", bold=True, size=12, color=accent)
                title_cell.fill = PatternFill("solid", fgColor=dark)
                title_cell.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[1].height = 22

                # Infos filtres
                ws.merge_cells("A2:G2")
                info_parts = []
                if sexe:  info_parts.append(f"Sexe : {sexe}")
                if etab:  info_parts.append(f"Établissement : {etab}")
                if cls:   info_parts.append(f"Classe : {cls}")
                info_parts.append(f"{len(rows)} participant(s)")
                ws["A2"].value = "  |  ".join(info_parts)
                ws["A2"].font  = Font(name="Arial", size=8, color="FF555D72")
                ws["A2"].fill  = PatternFill("solid", fgColor=dark)
                ws.row_dimensions[2].height = 14

                # En-têtes
                headers = ["Dossard", "Nom", "Prénom", "Classe", "Établissement", "Sexe", "VMA"]
                col_w   = [10, 20, 20, 12, 25, 8, 8]
                for ci, (h, w) in enumerate(zip(headers, col_w), 1):
                    cell = ws.cell(row=3, column=ci, value=h)
                    cell.font      = Font(name="Arial", bold=True, size=9, color=accent)
                    cell.fill      = PatternFill("solid", fgColor=dark)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = thin
                    ws.column_dimensions[get_column_letter(ci)].width = w
                ws.row_dimensions[3].height = 18

                # Données
                for ri, r in enumerate(rows, 4):
                    bg = light if ri % 2 == 0 else "FFFFFFFF"
                    vals = [
                        r.get("dossard") or "—",
                        r.get("nom") or "—",
                        r.get("prenom") or "—",
                        r.get("classe") or "—",
                        r.get("etablissement") or "—",
                        r.get("sexe") or "—",
                        r.get("vma") or "—",
                    ]
                    for ci, val in enumerate(vals, 1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        cell.font      = Font(name="Arial", size=9, color="FF1A1E2E")
                        cell.fill      = PatternFill("solid", fgColor=bg)
                        cell.alignment = Alignment(horizontal="center" if ci in (1,6,7) else "left", vertical="center")
                        cell.border    = thin
                    ws.row_dimensions[ri].height = 16

                ws.freeze_panes = "A4"
                path = os.path.join(exports_dir, f"{safe_title}.xlsx")
                wb.save(path)

            # ════════════════════════════════
            else:  # pdf
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import ParagraphStyle

                path = os.path.join(exports_dir, f"{safe_title}.pdf")
                doc  = SimpleDocTemplate(path, pagesize=A4,
                    leftMargin=15*mm, rightMargin=15*mm,
                    topMargin=15*mm, bottomMargin=15*mm)

                story = []
                story.append(Paragraph("CROSSCONTROL", ParagraphStyle("t", fontSize=16,
                    fontName="Helvetica-Bold", spaceAfter=1*mm, textColor=colors.HexColor("#d60a3c"))))
                story.append(Paragraph(f"{title_line} — {event.get('nom','')}", ParagraphStyle("s",
                    fontSize=9, fontName="Helvetica", spaceAfter=2*mm, textColor=colors.HexColor("#555d72"))))

                meta = []
                if sexe:  meta.append(f"Sexe : {sexe}")
                if etab:  meta.append(f"Établissement : {etab}")
                if cls:   meta.append(f"Classe : {cls}")
                meta.append(f"{len(rows)} participant(s)")
                story.append(Paragraph("  |  ".join(meta), ParagraphStyle("i",
                    fontSize=8, fontName="Helvetica", spaceAfter=5*mm, textColor=colors.HexColor("#333"))))

                if context == "course" and course:
                    c = dict(course)
                    course_meta = []
                    if c.get("distance"): course_meta.append(f"Distance : {c['distance']}m")
                    if c.get("vma_min") and c.get("vma_max"): course_meta.append(f"VMA : {c['vma_min']}–{c['vma_max']} km/h")
                    if course_meta:
                        story.append(Paragraph("  |  ".join(course_meta), ParagraphStyle("cm",
                            fontSize=8, fontName="Helvetica", spaceAfter=5*mm, textColor=colors.HexColor("#333"))))

                headers = ["Dossard", "Nom", "Prénom", "Classe", "Établissement", "Sexe", "VMA"]
                data = [headers] + [[
                    str(r.get("dossard") or "—"), r.get("nom") or "—", r.get("prenom") or "—",
                    r.get("classe") or "—", r.get("etablissement") or "—",
                    r.get("sexe") or "—", str(r.get("vma") or "—")
                ] for r in rows]

                col_widths = [18*mm, 35*mm, 35*mm, 20*mm, 43*mm, 12*mm, 12*mm]
                t = Table(data, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0), (-1,0),  colors.HexColor("#161920")),
                    ("TEXTCOLOR",     (0,0), (-1,0),  colors.HexColor("#d60a3c")),
                    ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
                    ("FONTSIZE",      (0,0), (-1,0),  8),
                    ("TOPPADDING",    (0,0), (-1,0),  3*mm),
                    ("BOTTOMPADDING", (0,0), (-1,0),  3*mm),
                    ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
                    ("FONTSIZE",      (0,1), (-1,-1), 8),
                    ("TOPPADDING",    (0,1), (-1,-1), 2.5*mm),
                    ("BOTTOMPADDING", (0,1), (-1,-1), 2.5*mm),
                    ("ROWBACKGROUNDS",(0,1), (-1,-1),  [colors.HexColor("#f7f7f7"), colors.white]),
                    ("TEXTCOLOR",     (0,1), (-1,-1), colors.HexColor("#1a1e2e")),
                    ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#d0d5e2")),
                    ("LINEBELOW",     (0,0), (-1,0),  1, colors.HexColor("#d60a3c")),
                    ("ALIGN",         (0,0), (0,-1),  "CENTER"),
                    ("ALIGN",         (5,0), (6,-1),  "CENTER"),
                ]))
                story.append(t)
                doc.build(story)

            # ── Ouvrir le fichier ──
            import subprocess
            if os.name == "nt":
                os.startfile(path)
            else:
                subprocess.Popen(["xdg-open", path])

            return {"success": True, "path": path}

        except Exception as e:
            log.error(f"export_file ERREUR: {e}\n{traceback.format_exc()}")
            return {"success": False, "error": str(e)}

    def get_stats(self):
        try:
            conn = get_event_db()
            return {
                "participants": conn.execute("SELECT COUNT(*) as c FROM participants").fetchone()['c'],
                "courses":      conn.execute("SELECT COUNT(*) as c FROM courses").fetchone()['c'],
                "arrivees":     conn.execute("SELECT COUNT(*) as c FROM arrivees WHERE participant_id IS NOT NULL").fetchone()['c'],
            }
        except Exception:
            return {"participants": 0, "courses": 0, "arrivees": 0}


# ════════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════════
def main():
    log.info("=== Démarrage CrossControl ===")
    log.info(f"Python {sys.version}")
    init_events_db()
    html_path = os.path.join(BASE_DIR, "app.html")
    webview.create_window(
        "CrossControl",
        html_path,
        js_api=API(),
        width=1920,
        height=1080,
        min_size=(900, 600)
    )
    webview.start(debug=True)

if __name__ == "__main__":
    main()